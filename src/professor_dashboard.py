"""
Professor Attendance Dashboard
Web-based interface for scanning IDs and exporting reports

Run with: python professor_dashboard.py
Then open: http://localhost:5000
"""

from flask import Flask, render_template, Response, jsonify, send_file, request
import cv2
import numpy as np
import pytesseract
from datetime import datetime
import sqlite3
import re
import time
from pathlib import Path
import pickle
from collections import Counter
import io
import os

# For exports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️ PDF export unavailable. Install: pip install reportlab")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Excel export unavailable. Install: pip install openpyxl")


app = Flask(__name__)


class AttendanceSystem:
    """Core attendance tracking system with camera and OCR"""
    
    def __init__(self):
        self.course_id = "MSC_CS"
        self.course_name = "MSc Computer Science"
        self.db_path = "attendance.db"
        self.is_scanning = False
        self.camera = None
        self.last_detected = {}
        self.cooldown = 4
        self.recent_scan = None
        self.recent_scan_time = 0
        
        # Detection state
        self.detection_history = []
        self.history_size = 5
        self.lost_frame_count = 0
        self.state = "waiting"
        self.detection_start = 0
        self.detection_confidence = 0
        
        # Load trained model
        self.model_loaded = self._load_model()
        
        self._init_database()
    
    def _init_database(self):
        """Initialize SQLite database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                name TEXT,
                last_seen DATETIME
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL,
                course_id TEXT NOT NULL,
                date DATE NOT NULL,
                timestamp DATETIME NOT NULL,
                status TEXT DEFAULT 'present'
            )
        """)
        conn.commit()
        conn.close()
    
    def _load_model(self) -> bool:
        """Load trained ID card detection model"""
        model_path = Path("models/enhanced_id_model.npz")
        if not model_path.exists():
            return False
        try:
            data = np.load(model_path)
            self.avg_color_hist = data['avg_color_hist']
            self.std_color_hist = data['std_color_hist']
            self.avg_aspect_ratio = float(data['avg_aspect_ratio'])
            self.std_aspect_ratio = float(data['std_aspect_ratio'])
            self.composite_template = data['composite_template']
            
            templates_path = Path("models/templates.pkl")
            if templates_path.exists():
                with open(templates_path, 'rb') as f:
                    self.templates = pickle.load(f)
            else:
                self.templates = [self.composite_template]
            return True
        except Exception as e:
            print(f"Model load error: {e}")
            return False
    
    def start_camera(self):
        """Start camera for scanning"""
        if self.camera is None:
            self.camera = cv2.VideoCapture(0)
            self.camera.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
            self.camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        self.is_scanning = True
        self.state = "waiting"
        self.detection_history = []
    
    def stop_camera(self):
        """Stop camera"""
        self.is_scanning = False
        if self.camera:
            self.camera.release()
            self.camera = None
    
    def get_attendance_today(self):
        """Get all attendance records for today"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT a.student_id, s.name, a.timestamp
            FROM attendance a
            LEFT JOIN students s ON a.student_id = s.student_id
            WHERE a.course_id = ? AND a.date = ?
            ORDER BY a.timestamp
        """, (self.course_id, today))
        records = cursor.fetchall()
        conn.close()
        return records
    
    def mark_attendance(self, student_id, name):
        """Mark a student as present"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        now = datetime.now()
        today = now.strftime('%Y-%m-%d')
        timestamp = now.isoformat()
        
        cursor.execute("""
            INSERT OR REPLACE INTO students (student_id, name, last_seen)
            VALUES (?, ?, ?)
        """, (student_id, name, timestamp))
        
        cursor.execute("""
            SELECT COUNT(*) FROM attendance 
            WHERE student_id = ? AND course_id = ? AND date = ?
        """, (student_id, self.course_id, today))
        
        if cursor.fetchone()[0] == 0:
            cursor.execute("""
                INSERT INTO attendance (student_id, course_id, date, timestamp, status)
                VALUES (?, ?, ?, ?, 'present')
            """, (student_id, self.course_id, today, timestamp))
            conn.commit()
            conn.close()
            return True
        
        conn.close()
        return False
    
    def clear_today(self):
        """Clear today's attendance records"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("DELETE FROM attendance WHERE course_id = ? AND date = ?", 
                      (self.course_id, today))
        conn.commit()
        conn.close()
        self.last_detected = {}
    
    def _vote_on_digits(self, id_list):
        """Vote on each digit position for accuracy"""
        if not id_list:
            return None
        valid_ids = [i for i in id_list if len(i) == 8 and i.isdigit()]
        if not valid_ids:
            return id_list[0] if id_list else None
        
        result = []
        for pos in range(8):
            digits_at_pos = [i[pos] for i in valid_ids]
            digit_counts = Counter(digits_at_pos)
            result.append(digit_counts.most_common(1)[0][0])
        return ''.join(result)
    
    def extract_text(self, roi):
        """Extract text from ID card region using OCR"""
        height, width = roi.shape[:2]
        scale = max(1.0, 800 / width)
        if scale != 1.0:
            roi = cv2.resize(roi, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        
        results = []
        ids_found = []
        
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                       cv2.THRESH_BINARY, 11, 2)
        
        for psm in ['--psm 11', '--psm 6', '--psm 4']:
            try:
                text = pytesseract.image_to_string(thresh, config=psm)
                results.append(text)
                ids = re.findall(r'\b(\d{8})\b', text)
                ids_found.extend(ids)
            except:
                pass
        
        if ids_found:
            id_counts = Counter(ids_found)
            best_id = max(id_counts, key=id_counts.get)
            for result in results:
                if best_id in result:
                    return result
        
        return max(results, key=len) if results else ""
    
    def parse_id(self, text):
        """Parse student ID from OCR text (name detection disabled)"""
        student_id = None
        
        all_ids = re.findall(r'\b(\d{8})\b', text)
        
        def is_likely_date(candidate):
            """Check if 8-digit number looks like a date"""
            first_two = int(candidate[:2])
            middle_two = int(candidate[2:4])
            last_four = int(candidate[4:])
            
            # DDMMYYYY format (e.g., 02102002 = 2nd Oct 2002)
            if (1 <= first_two <= 31 and 1 <= middle_two <= 12 and 1980 <= last_four <= 2030):
                return True
            # MMDDYYYY format
            if (1 <= first_two <= 12 and 1 <= middle_two <= 31 and 1980 <= last_four <= 2030):
                return True
            return False
        
        # First priority: Maynooth IDs start with '2' (e.g., 24250414)
        for candidate in all_ids:
            if candidate.startswith('2') and not is_likely_date(candidate):
                student_id = candidate
                break
        
        # Fallback: any 8-digit number that's NOT a date
        if not student_id:
            for candidate in all_ids:
                if not is_likely_date(candidate):
                    student_id = candidate
                    break
        
        if student_id:
            return {'student_id': student_id, 'name': 'Student'}
        return None
    
    def detect_card(self, frame, roi_bounds):
        """Detect if an ID card is present in the region"""
        x, y, w, h = roi_bounds
        roi = frame[y:y+h, x:x+w]
        
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 50, 150)
        
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        for contour in contours:
            cx, cy, cw, ch = cv2.boundingRect(contour)
            area = cw * ch
            aspect = cw / float(ch) if ch > 0 else 0
            
            if cw > 150 and ch > 80 and area > 20000 and 1.3 < aspect < 2.0:
                self.detection_confidence = 0.7
                return True
        
        return False
    
    def process_frame(self, frame):
        """Process a camera frame"""
        if not self.is_scanning:
            return frame
        
        height, width = frame.shape[:2]
        current_time = time.time()
        
        # Card region
        card_w = int(width * 0.55)
        card_h = int(card_w / 1.6)
        card_x = (width - card_w) // 2
        card_y = (height - card_h) // 2 - 30
        card_bounds = (card_x, card_y, card_w, card_h)
        
        # Darken outside
        mask = np.zeros((height, width), dtype=np.uint8)
        cv2.rectangle(mask, (card_x, card_y), (card_x + card_w, card_y + card_h), 255, -1)
        dark = np.zeros_like(frame)
        darkened = cv2.addWeighted(frame, 0.4, dark, 0.6, 0)
        mask_3ch = cv2.merge([mask, mask, mask])
        frame = np.where(mask_3ch == 255, frame, darkened)
        
        # Colors
        colors = {
            "waiting": (255, 255, 255),
            "detected": (0, 165, 255),
            "scanning": (0, 255, 255),
            "success": (0, 255, 0)
        }
        color = colors.get(self.state, (255, 255, 255))
        
        # Draw corners
        corner_len = 40
        for (cx, cy) in [(card_x, card_y), (card_x + card_w, card_y), 
                          (card_x, card_y + card_h), (card_x + card_w, card_y + card_h)]:
            dx = corner_len if cx == card_x else -corner_len
            dy = corner_len if cy == card_y else -corner_len
            cv2.line(frame, (cx, cy), (cx + dx, cy), color, 3)
            cv2.line(frame, (cx, cy), (cx, cy + dy), color, 3)
        
        # State machine
        if self.state == "waiting":
            detected = self.detect_card(frame, card_bounds)
            if detected:
                self.detection_history.append(True)
                if len(self.detection_history) > self.history_size:
                    self.detection_history.pop(0)
                if sum(self.detection_history) >= 2:
                    self.state = "detected"
                    self.detection_start = current_time
                    self.lost_frame_count = 0
            else:
                self.detection_history.append(False)
                if len(self.detection_history) > self.history_size:
                    self.detection_history.pop(0)
                    
        elif self.state == "detected":
            detected = self.detect_card(frame, card_bounds)
            if detected:
                self.lost_frame_count = 0
            else:
                self.lost_frame_count += 1
            
            if self.lost_frame_count > 8:
                self.state = "waiting"
                self.detection_history = []
            elif current_time - self.detection_start > 1.0:
                self.state = "scanning"
                
        elif self.state == "scanning":
            x, y, w, h = card_bounds
            
            all_ids = []
            
            for _ in range(5):
                ret, scan_frame = self.camera.read()
                if ret:
                    roi = scan_frame[y:y+h, x:x+w]
                    text = self.extract_text(roi)
                    student = self.parse_id(text)
                    if student:
                        all_ids.append(student['student_id'])
                time.sleep(0.03)
            
            final_id = self._vote_on_digits(all_ids)
            
            if final_id:
                last = self.last_detected.get(final_id, 0)
                if current_time - last > self.cooldown:
                    if self.mark_attendance(final_id, "Student"):
                        self.recent_scan = {'id': final_id, 'name': final_id, 'new': True}
                    else:
                        self.recent_scan = {'id': final_id, 'name': final_id, 'new': False}
                    self.recent_scan_time = current_time
                    self.last_detected[final_id] = current_time
                    self.state = "success"
                else:
                    self.state = "waiting"
            else:
                self.state = "waiting"
            
            self.detection_history = []
            
        elif self.state == "success":
            if current_time - self.recent_scan_time > 2.0:
                self.state = "waiting"
        
        # Status text
        status_text = {
            "waiting": "Hold ID card in frame",
            "detected": "Card detected - hold still...",
            "scanning": "Scanning...",
            "success": f"Marked: {self.recent_scan['name'] if self.recent_scan else ''}"
        }.get(self.state, "Ready")
        
        cv2.putText(frame, status_text, (card_x, card_y - 20),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
        
        return frame
    
    def generate_frames(self):
        """Generator for video streaming"""
        while self.is_scanning:
            if self.camera is None:
                break
            
            ret, frame = self.camera.read()
            if not ret:
                break
            
            frame = self.process_frame(frame)
            
            ret, buffer = cv2.imencode('.jpg', frame)
            if ret:
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')


# Initialize
system = AttendanceSystem()


# Routes
@app.route('/')
def index():
    return render_template('dashboard.html')


@app.route('/start_scanning', methods=['POST'])
def start_scanning():
    system.start_camera()
    return jsonify({'success': True})


@app.route('/stop_scanning', methods=['POST'])
def stop_scanning():
    system.stop_camera()
    return jsonify({'success': True})


@app.route('/video_feed')
def video_feed():
    return Response(system.generate_frames(),
                   mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/attendance')
def get_attendance():
    records = system.get_attendance_today()
    formatted = []
    for sid, name, timestamp in records:
        time_str = datetime.fromisoformat(timestamp).strftime('%H:%M:%S')
        formatted.append({
            'student_id': sid,
            'name': name or 'Unknown',
            'time': time_str
        })
    return jsonify({'records': formatted})


@app.route('/clear_today', methods=['POST'])
def clear_today():
    system.clear_today()
    return jsonify({'success': True})


@app.route('/export/pdf')
def export_pdf():
    # Stop camera before export to prevent crash
    system.stop_camera()
    
    if not PDF_AVAILABLE:
        return "PDF export unavailable. Install: pip install reportlab", 500
    
    records = system.get_attendance_today()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, 
                                  spaceAfter=30, alignment=1)
    elements.append(Paragraph("Attendance Report", title_style))
    
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=12, spaceAfter=6)
    elements.append(Paragraph(f"<b>Course:</b> {system.course_name}", info_style))
    elements.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%A, %d %B %Y')}", info_style))
    elements.append(Paragraph(f"<b>Total Students:</b> {len(records)}", info_style))
    elements.append(Spacer(1, 20))
    
    data = [['#', 'Student ID', 'Name', 'Time']]
    for i, (sid, name, timestamp) in enumerate(records, 1):
        time_str = datetime.fromisoformat(timestamp).strftime('%H:%M:%S')
        data.append([str(i), sid, name or 'Unknown', time_str])
    
    table = Table(data, colWidths=[0.5*inch, 1.3*inch, 2.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
    ]))
    elements.append(table)
    
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, 
                                   textColor=colors.grey)
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Maynooth University",
        footer_style
    ))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"attendance_{system.course_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/export/excel')
def export_excel():
    # Stop camera before export to prevent crash
    system.stop_camera()
    
    if not EXCEL_AVAILABLE:
        return "Excel export unavailable. Install: pip install openpyxl", 500
    
    records = system.get_attendance_today()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Attendance - {system.course_name} - {datetime.now().strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['#', 'Student ID', 'Name', 'Time']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for i, (sid, name, timestamp) in enumerate(records, 1):
        time_str = datetime.fromisoformat(timestamp).strftime('%H:%M:%S')
        row = 3 + i
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=sid).border = border
        ws.cell(row=row, column=3, value=name or 'Unknown').border = border
        ws.cell(row=row, column=4, value=time_str).border = border
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"attendance_{system.course_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    print("\n" + "="*60)
    print("🎓 PROFESSOR ATTENDANCE DASHBOARD")
    print("="*60)
    print(f"\n📍 Open in browser: http://localhost:5000")
    print(f"📚 Course: {system.course_name}")
    print("\nPress Ctrl+C to stop the server")
    print("="*60 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)