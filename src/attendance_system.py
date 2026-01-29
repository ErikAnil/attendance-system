"""
Student Attendance System - Video Processing Backend
For macOS - Processes videos of student IDs and extracts attendance data
"""

import cv2
import pytesseract
from pathlib import Path
import re
from datetime import datetime
import sqlite3
import json
from typing import List, Dict, Optional
import numpy as np

class AttendanceSystem:
    def __init__(self, db_path: str = "attendance.db"):
        self.db_path = db_path
        self.init_database()
        
    def init_database(self):
        """Initialize SQLite database with required tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Students table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                email TEXT,
                department TEXT,
                photo_path TEXT
            )
        """)
        
        # Courses table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS courses (
                course_id TEXT PRIMARY KEY,
                course_name TEXT NOT NULL,
                professor_name TEXT,
                semester TEXT
            )
        """)
        
        # Attendance records
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL,
                course_id TEXT NOT NULL,
                date DATE NOT NULL,
                timestamp DATETIME NOT NULL,
                confidence REAL,
                status TEXT DEFAULT 'present',
                FOREIGN KEY (student_id) REFERENCES students(student_id),
                FOREIGN KEY (course_id) REFERENCES courses(course_id),
                UNIQUE(student_id, course_id, date)
            )
        """)
        
        # Processing logs
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS processing_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                video_path TEXT,
                course_id TEXT,
                date DATE,
                processed_at DATETIME,
                total_students INTEGER,
                status TEXT
            )
        """)
        
        conn.commit()
        conn.close()
        print("✓ Database initialized")
    
    def extract_frames(self, video_path: str, frame_interval: int = 30) -> List[np.ndarray]:
        """
        Extract frames from video at specified intervals
        
        Args:
            video_path: Path to video file
            frame_interval: Extract every Nth frame (default 30 = 1 per second at 30fps)
        """
        frames = []
        cap = cv2.VideoCapture(video_path)
        
        if not cap.isOpened():
            raise ValueError(f"Cannot open video: {video_path}")
        
        fps = cap.get(cv2.CAP_PROP_FPS)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        
        print(f"Video: {fps:.1f} FPS, {total_frames} frames")
        
        frame_count = 0
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
            
            # Extract frame at intervals
            if frame_count % frame_interval == 0:
                # Check if frame is not too blurry
                if self.is_frame_clear(frame):
                    frames.append(frame)
            
            frame_count += 1
        
        cap.release()
        print(f"✓ Extracted {len(frames)} clear frames")
        return frames
    
    def is_frame_clear(self, frame: np.ndarray, threshold: float = 100.0) -> bool:
        """
        Check if frame is clear enough (not blurry) using Laplacian variance
        
        Args:
            frame: Input frame
            threshold: Minimum variance threshold (higher = clearer)
        """
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        return laplacian_var > threshold
    
    def detect_id_card(self, frame: np.ndarray, use_trained_model: bool = True) -> Optional[np.ndarray]:
        """
        Detect and extract ID card region from frame
        Uses custom trained model if available, otherwise falls back to basic detection
        """
        # Try to use custom trained detector
        if use_trained_model:
            try:
                from id_card_detector import IDCardDetector
                detector = IDCardDetector()
                detections = detector.detect_id_cards(frame, threshold=0.4)
                
                if detections:
                    # Get highest confidence detection
                    x, y, w, h, conf = max(detections, key=lambda d: d[4])
                    return frame[y:y+h, x:x+w]
            except:
                pass  # Fall back to basic detection
        
        # Fallback: Basic contour detection
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 50, 150)
        
        # Find contours
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # Find largest rectangular contour (likely the ID card)
        for contour in sorted(contours, key=cv2.contourArea, reverse=True)[:5]:
            peri = cv2.arcLength(contour, True)
            approx = cv2.approxPolyDP(contour, 0.02 * peri, True)
            
            # If contour has 4 points, it's likely rectangular (ID card)
            if len(approx) == 4:
                x, y, w, h = cv2.boundingRect(approx)
                
                # Check if aspect ratio is reasonable for ID card
                aspect_ratio = w / float(h)
                if 1.3 < aspect_ratio < 2.0 and w > 200 and h > 100:
                    return frame[y:y+h, x:x+w]
        
        # Final fallback: use center region
        height, width = frame.shape[:2]
        y1, y2 = height//4, 3*height//4
        x1, x2 = width//4, 3*width//4
        return frame[y1:y2, x1:x2]
    
    def preprocess_image(self, image: np.ndarray) -> np.ndarray:
        """
        Preprocess image for better OCR results
        """
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Apply adaptive thresholding
        thresh = cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
            cv2.THRESH_BINARY, 11, 2
        )
        
        # Denoise
        denoised = cv2.fastNlMeansDenoising(thresh)
        
        return denoised
    
    def extract_text(self, image: np.ndarray) -> str:
        """
        Extract text from image using Tesseract OCR
        """
        preprocessed = self.preprocess_image(image)
        
        # Configure Tesseract
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(preprocessed, config=custom_config)
        
        return text
    
    def parse_student_info(self, text: str) -> Optional[Dict[str, str]]:
        """
        Parse student information from OCR text
        Adjust regex patterns based on your institution's ID format
        """
        # Common patterns - adjust based on your ID card format
        patterns = {
            'student_id': [
                r'(?:ID|Student ID|Roll|Reg)[:\s#]*([A-Z0-9\-]+)',
                r'\b([A-Z]{2,3}\d{6,10})\b',  # Pattern like CS2023001
                r'\b(\d{8,10})\b'  # Pure numeric IDs
            ],
            'name': [
                r'(?:Name|Student)[:\s]*([A-Z][a-z]+(?:\s[A-Z][a-z]+)+)',
                r'\n([A-Z][a-z]+\s+[A-Z][a-z]+)\n'
            ],
            'email': [
                r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
            ],
            'department': [
                r'(?:Dept|Department|Program)[:\s]*([A-Z][a-zA-Z\s&]+)'
            ]
        }
        
        result = {}
        
        # Try each pattern until we find a match
        for field, pattern_list in patterns.items():
            for pattern in pattern_list:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    result[field] = match.group(1).strip()
                    break
        
        # Must have at least student_id to be valid
        if 'student_id' in result:
            return result
        
        return None
    
    def process_video(self, video_path: str, course_id: str, date: str = None) -> Dict:
        """
        Main processing pipeline for video
        
        Args:
            video_path: Path to video file
            course_id: Course identifier
            date: Date of attendance (YYYY-MM-DD), defaults to today
        
        Returns:
            Dictionary with processing results
        """
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
        
        print(f"\n{'='*60}")
        print(f"Processing Attendance Video")
        print(f"Course: {course_id} | Date: {date}")
        print(f"{'='*60}\n")
        
        # Extract frames
        frames = self.extract_frames(video_path)
        
        # Process each frame
        students_found = []
        seen_ids = set()
        
        print("Processing frames...")
        for i, frame in enumerate(frames):
            print(f"Frame {i+1}/{len(frames)}", end='\r')
            
            # Detect ID card region
            id_region = self.detect_id_card(frame)
            if id_region is None:
                continue
            
            # Extract text
            text = self.extract_text(id_region)
            
            # Parse student info
            student_info = self.parse_student_info(text)
            
            if student_info and student_info['student_id'] not in seen_ids:
                students_found.append(student_info)
                seen_ids.add(student_info['student_id'])
                print(f"\n✓ Found: {student_info['student_id']} - {student_info.get('name', 'Unknown')}")
        
        print(f"\n\nTotal students identified: {len(students_found)}")
        
        # Save to database
        self.save_attendance(students_found, course_id, date)
        
        # Log processing
        self.log_processing(video_path, course_id, date, len(students_found))
        
        return {
            'success': True,
            'students_count': len(students_found),
            'students': students_found,
            'course_id': course_id,
            'date': date
        }
    
    def save_attendance(self, students: List[Dict], course_id: str, date: str):
        """Save attendance records to database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        timestamp = datetime.now().isoformat()
        
        for student in students:
            student_id = student['student_id']
            name = student.get('name', 'Unknown')
            
            # Insert/update student record
            cursor.execute("""
                INSERT OR REPLACE INTO students (student_id, name, email, department)
                VALUES (?, ?, ?, ?)
            """, (
                student_id,
                name,
                student.get('email'),
                student.get('department')
            ))
            
            # Insert attendance record
            try:
                cursor.execute("""
                    INSERT INTO attendance (student_id, course_id, date, timestamp, status)
                    VALUES (?, ?, ?, ?, 'present')
                """, (student_id, course_id, date, timestamp))
            except sqlite3.IntegrityError:
                # Already marked present for this date
                pass
        
        conn.commit()
        conn.close()
        print("✓ Attendance saved to database")
    
    def log_processing(self, video_path: str, course_id: str, date: str, count: int):
        """Log video processing details"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO processing_logs 
            (video_path, course_id, date, processed_at, total_students, status)
            VALUES (?, ?, ?, ?, ?, 'completed')
        """, (video_path, course_id, date, datetime.now().isoformat(), count))
        
        conn.commit()
        conn.close()
    
    def get_attendance_report(self, course_id: str, date: str = None) -> List[Dict]:
        """Generate attendance report for a course"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = """
            SELECT s.student_id, s.name, s.department, a.timestamp, a.status
            FROM attendance a
            JOIN students s ON a.student_id = s.student_id
            WHERE a.course_id = ?
        """
        
        params = [course_id]
        if date:
            query += " AND a.date = ?"
            params.append(date)
        
        query += " ORDER BY s.student_id"
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()
        
        return [
            {
                'student_id': r[0],
                'name': r[1],
                'department': r[2],
                'timestamp': r[3],
                'status': r[4]
            }
            for r in results
        ]
    
    def export_to_csv(self, course_id: str, date: str, output_path: str):
        """Export attendance to CSV"""
        import csv
        
        records = self.get_attendance_report(course_id, date)
        
        with open(output_path, 'w', newline='') as f:
            if records:
                writer = csv.DictWriter(f, fieldnames=records[0].keys())
                writer.writeheader()
                writer.writerows(records)
        
        print(f"✓ Exported to {output_path}")
    
    def export_to_excel(self, course_id: str, date: str, output_path: str):
        """Export attendance to Excel with formatting"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
        except ImportError:
            print("⚠ openpyxl not installed. Installing now...")
            import subprocess
            subprocess.check_call(['pip', 'install', 'openpyxl'])
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        records = self.get_attendance_report(course_id, date)
        
        if not records:
            print("No records to export")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Add title
        ws['A1'] = f"Attendance Report - {course_id}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Date: {date if date else 'All Dates'}"
        ws['A2'].font = Font(size=12)
        
        # Headers
        headers = ['Student ID', 'Name', 'Department', 'Date', 'Time', 'Status']
        header_row = 4
        
        # Style for headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, record in enumerate(records, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=record['student_id']).border = border
            ws.cell(row=row_idx, column=2, value=record['name']).border = border
            ws.cell(row=row_idx, column=3, value=record['department']).border = border
            ws.cell(row=row_idx, column=4, value=record['timestamp'].split('T')[0]).border = border
            ws.cell(row=row_idx, column=5, value=record['timestamp'].split('T')[1].split('.')[0]).border = border
            
            status_cell = ws.cell(row=row_idx, column=6, value=record['status'].upper())
            status_cell.border = border
            
            # Color code status
            if record['status'] == 'present':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                status_cell.font = Font(color='006100', bold=True)
            
            status_cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary
        summary_row = len(records) + header_row + 2
        ws[f'A{summary_row}'] = f"Total Students: {len(records)}"
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        wb.save(output_path)
        print(f"✓ Exported to {output_path}")
    
    def export_to_pdf(self, course_id: str, date: str, output_path: str):
        """Export attendance to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from datetime import datetime
        except ImportError:
            print("⚠ reportlab not installed. Installing now...")
            import subprocess
            subprocess.check_call(['pip', 'install', 'reportlab'])
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        
        records = self.get_attendance_report(course_id, date)
        
        if not records:
            print("No records to export")
            return
        
        # Create PDF
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=1
        )
        
        # Title
        title = Paragraph(f"Attendance Report", title_style)
        elements.append(title)
        
        # Subtitle
        subtitle = Paragraph(
            f"Course: {course_id} | Date: {date if date else 'All Dates'} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            subtitle_style
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        data = [['Student ID', 'Name', 'Department', 'Date', 'Time', 'Status']]
        
        for record in records:
            data.append([
                record['student_id'],
                record['name'],
                record['department'] or 'N/A',
                record['timestamp'].split('T')[0],
                record['timestamp'].split('T')[1].split('.')[0],
                record['status'].upper()
            ])
        
        # Create table
        table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.3*inch, 1*inch, 0.9*inch, 0.8*inch])
        
        # Table style
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            
            # Status column styling
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),
            ('FONTNAME', (5, 1), (5, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Summary
        elements.append(Spacer(1, 0.3*inch))
        summary = Paragraph(
            f"<b>Total Students Present: {len(records)}</b>",
            styles['Normal']
        )
        elements.append(summary)
        
        # Build PDF
        doc.build(elements)
        print(f"✓ Exported to {output_path}")


def main():
    """Example usage"""
    import sys
    
    # Initialize system
    system = AttendanceSystem()
    
    # Check for export commands
    if len(sys.argv) > 1 and sys.argv[1] == '--export':
        if len(sys.argv) < 5:
            print("Usage for export:")
            print("  python attendance_system.py --export <format> <course_id> <date>")
            print("\nFormats: csv, excel, pdf")
            print("Example: python attendance_system.py --export excel CS501 2024-12-15")
            return
        
        export_format = sys.argv[2].lower()
        course_id = sys.argv[3]
        date = sys.argv[4]
        
        output_file = f"attendance_{course_id}_{date}.{export_format if export_format != 'excel' else 'xlsx'}"
        
        if export_format == 'csv':
            system.export_to_csv(course_id, date, output_file)
        elif export_format == 'excel':
            system.export_to_excel(course_id, date, output_file)
        elif export_format == 'pdf':
            system.export_to_pdf(course_id, date, output_file)
        else:
            print(f"Unknown format: {export_format}")
            print("Available formats: csv, excel, pdf")
        
        return
    
    # Example: Process a video
    if len(sys.argv) > 1:
        video_path = sys.argv[1]
        course_id = sys.argv[2] if len(sys.argv) > 2 else "CS501"
        
        result = system.process_video(video_path, course_id)
        
        print(f"\n{'='*60}")
        print("PROCESSING COMPLETE")
        print(f"{'='*60}")
        print(f"Students processed: {result['students_count']}")
        print(f"Course: {result['course_id']}")
        print(f"Date: {result['date']}")
        
        # Generate report
        print("\nAttendance Report:")
        print("-" * 60)
        report = system.get_attendance_report(course_id, result['date'])
        for record in report:
            print(f"{record['student_id']:15} {record['name']:25} {record['status']}")
        
        # Ask if user wants to export
        print("\n" + "="*60)
        print("Export options:")
        print(f"  CSV:   python attendance_system.py --export csv {course_id} {result['date']}")
        print(f"  Excel: python attendance_system.py --export excel {course_id} {result['date']}")
        print(f"  PDF:   python attendance_system.py --export pdf {course_id} {result['date']}")
        
    else:
        print("Usage: python attendance_system.py <video_path> [course_id]")
        print("\nExample:")
        print("  python attendance_system.py ~/Videos/class_2024_12_14.mov CS501")
        print("\nExport usage:")
        print("  python attendance_system.py --export excel CS501 2024-12-15")


if __name__ == "__main__":
    main()